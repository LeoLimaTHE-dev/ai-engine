from pathlib import Path

from openpyxl import load_workbook

from ai_engine.models import (
    DocumentContent,
    DocumentTable,
)


def read_xlsx(
    file_path: str | Path,
) -> DocumentContent:
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    if path.suffix.lower() not in (
        ".xlsx",
        ".xlsm",
    ):
        raise ValueError(f"Expected an .xlsx or .xlsm file, got: {path.suffix}")

    workbook = load_workbook(
        filename=path,
        data_only=True,
        read_only=False,
    )

    tables: list[DocumentTable] = []

    sheet_names: list[str] = []

    sheet_row_counts: dict[str, int] = {}

    sheet_column_counts: dict[str, int] = {}

    for worksheet in workbook.worksheets:
        sheet_names.append(worksheet.title)

        rows: list[list[str]] = []

        for row in worksheet.iter_rows(values_only=True):
            values = []

            for value in row:
                if value is None:
                    values.append("")
                else:
                    values.append(str(value))

            # Ignore completely empty rows
            if any(value.strip() for value in values):
                rows.append(values)

        tables.append(
            DocumentTable(
                rows=rows,
                name=worksheet.title,
                source=worksheet.title,
            )
        )

        sheet_row_counts[worksheet.title] = len(rows)

        sheet_column_counts[worksheet.title] = max(
            (len(row) for row in rows),
            default=0,
        )

    workbook.close()

    return DocumentContent(
        source_path=path,
        tables=tables,
        metadata={
            "format": "xlsx",
            "filename": path.name,
            "sheet_count": len(sheet_names),
            "sheet_names": sheet_names,
            "sheet_row_counts": (sheet_row_counts),
            "sheet_column_counts": (sheet_column_counts),
        },
    )
