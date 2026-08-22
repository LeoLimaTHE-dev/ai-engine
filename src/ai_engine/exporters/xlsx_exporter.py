from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def save_xlsx(
    content: str,
    output_path: str | Path,
    title: str = "AI Result",
) -> Path:
    path = Path(output_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Resultado"

    worksheet["A1"] = title
    worksheet["A1"].font = Font(
        bold=True,
        size=14,
    )

    row_index = 3

    for line in content.splitlines():
        worksheet.cell(
            row=row_index,
            column=1,
            value=line,
        )

        worksheet.cell(
            row=row_index,
            column=1,
        ).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )

        row_index += 1

    worksheet.column_dimensions["A"].width = 100

    workbook.save(path)

    return path


def save_xlsx_tables(
    tables,
    output_path: str | Path,
) -> Path:
    path = Path(output_path)

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for table_index, table in enumerate(
        tables,
        start=1,
    ):
        sheet_name = table.name or f"Table {table_index}"

        # Excel limits worksheet names to 31 chars.
        sheet_name = sheet_name[:31]

        worksheet = workbook.create_sheet(title=sheet_name)

        row_index = 1

        if table.headers:
            for column_index, header in enumerate(
                table.headers,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=header,
                )

                cell.font = Font(bold=True)

            row_index += 1

        for row in table.rows:
            for column_index, value in enumerate(
                row,
                start=1,
            ):
                worksheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value,
                )

            row_index += 1

    if not workbook.sheetnames:
        workbook.create_sheet(title="Resultado")

    workbook.save(path)

    return path
