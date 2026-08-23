import pymupdf
import pytest
from docx import Document
from openpyxl import load_workbook

import ai_engine.actions as actions_module
from ai_engine.actions import execute_output, execute_structured_result
from ai_engine.results import OutputRequest, ResultTable, StructuredResult
from ai_engine.structured_errors import OutputExecutionError, OutputValidationError


def make_result(*outputs):
    return StructuredResult(message="Done", outputs=list(outputs))


def output(format="txt", filename="result.txt", content="content", **kwargs):
    return OutputRequest(
        format=format,
        filename=filename,
        content=content,
        **kwargs,
    )


def test_empty_result_creates_nothing_and_returns_empty_list(tmp_path):
    output_dir = tmp_path / "outputs"

    assert execute_structured_result(StructuredResult(message=""), output_dir) == []
    assert not output_dir.exists()


@pytest.mark.parametrize(
    ("format_name", "filename", "expected_content"),
    [
        ("txt", "result.txt", "plain"),
        ("md", "result.md", "# markdown"),
    ],
)
def test_text_outputs_use_planned_paths(
    format_name,
    filename,
    expected_content,
    tmp_path,
):
    created = execute_structured_result(
        make_result(output(format_name, filename, expected_content)),
        tmp_path,
    )

    assert created == [tmp_path / filename]
    assert created[0].read_text(encoding="utf-8") == expected_content


def test_docx_uses_planned_path_title_and_content(tmp_path):
    created = execute_structured_result(
        make_result(
            output(
                "docx",
                "report.docx",
                "Body",
                title="Title",
            )
        ),
        tmp_path,
    )[0]

    document = Document(created)
    assert [paragraph.text for paragraph in document.paragraphs] == ["Title", "Body"]


def test_pdf_uses_planned_path_title_and_content(tmp_path):
    created = execute_structured_result(
        make_result(output("pdf", "report.pdf", "Body", title="Title")),
        tmp_path,
    )[0]

    document = pymupdf.open(created)
    text = "\n".join(page.get_text() for page in document)
    document.close()
    assert "Title" in text
    assert "Body" in text


def test_xlsx_linear_uses_existing_linear_exporter(tmp_path):
    created = execute_structured_result(
        make_result(output("xlsx", "linear.xlsx", "First\nSecond", title="Title")),
        tmp_path,
    )[0]

    workbook = load_workbook(created)
    assert workbook.sheetnames == ["Resultado"]
    assert workbook["Resultado"]["A1"].value == "Title"
    assert workbook["Resultado"]["A3"].value == "First"
    workbook.close()


def test_xlsx_tabular_uses_planned_sheet_names(tmp_path):
    table = ResultTable(name="Data", headers=["A"], rows=[["value"]])
    created = execute_structured_result(
        make_result(output("xlsx", "tables.xlsx", tables=[table])),
        tmp_path,
    )[0]

    workbook = load_workbook(created)
    assert workbook.sheetnames == ["Data"]
    assert workbook["Data"]["A2"].value == "value"
    workbook.close()


def test_multiple_valid_outputs_preserve_planned_order(tmp_path):
    created = execute_structured_result(
        make_result(
            output("txt", "first.txt", "1"),
            output("md", "second.md", "2"),
            output("txt", "third.txt", "3"),
        ),
        tmp_path,
    )

    assert created == [
        tmp_path / "first.txt",
        tmp_path / "second.md",
        tmp_path / "third.txt",
    ]


def test_structured_execution_replaces_conflicting_extension(tmp_path):
    created = execute_structured_result(
        make_result(output("pdf", "relatorio.docx", "Body")),
        tmp_path,
    )

    assert created == [tmp_path / "relatorio.pdf"]
    assert not (tmp_path / "relatorio.docx.pdf").exists()


def test_direct_execute_output_keeps_historical_double_extension(tmp_path):
    created = execute_output(output("pdf", "relatorio.docx", "Body"), tmp_path)

    assert created == tmp_path / "relatorio.docx.pdf"


@pytest.mark.parametrize(
    "filename",
    ["../../portable.txt", r"..\..\portable.txt"],
)
def test_structured_execution_uses_portable_basename(filename, tmp_path):
    created = execute_structured_result(
        make_result(output("txt", filename, "Body")),
        tmp_path,
    )

    assert created == [tmp_path / "portable.txt"]


@pytest.mark.parametrize("filename", ["CON.txt", "bad:name.txt"])
def test_unsafe_filename_fails_before_output_directory_is_created(filename, tmp_path):
    output_dir = tmp_path / "outputs"

    with pytest.raises(OutputValidationError):
        execute_structured_result(
            make_result(output("txt", filename, "Body")),
            output_dir,
        )

    assert not output_dir.exists()


def test_output_collision_fails_before_any_file_is_written(tmp_path):
    output_dir = tmp_path / "outputs"
    result = make_result(
        output("txt", "a/report.txt", "A"),
        output("txt", "b/report.txt", "B"),
    )

    with pytest.raises(OutputValidationError, match="collides"):
        execute_structured_result(result, output_dir)

    assert not output_dir.exists()


def test_overwrite_false_rejects_existing_file_without_changing_it(tmp_path):
    existing = tmp_path / "result.txt"
    existing.write_text("old", encoding="utf-8")

    with pytest.raises(OutputValidationError, match="already exists"):
        execute_structured_result(
            make_result(output(content="new")),
            tmp_path,
            overwrite=False,
        )

    assert existing.read_text(encoding="utf-8") == "old"


@pytest.mark.parametrize(
    "invalid_output",
    [
        output("csv", "invalid.csv", "B"),
        output("txt", "CON.txt", "B"),
        output(
            "xlsx",
            "invalid.xlsx",
            tables=[
                ResultTable(
                    name="Data",
                    headers=["A", "B"],
                    rows=[["one"]],
                )
            ],
        ),
    ],
    ids=["format", "filename", "table"],
)
def test_all_structural_errors_are_detected_before_first_write(
    invalid_output,
    tmp_path,
):
    output_dir = tmp_path / "outputs"
    result = make_result(
        output("txt", "a.txt", "A"),
        invalid_output,
        output("txt", "c.txt", "C"),
    )

    with pytest.raises(OutputValidationError):
        execute_structured_result(result, output_dir)

    assert not output_dir.exists()
    assert not (output_dir / "a.txt").exists()
    assert not (output_dir / "c.txt").exists()


def test_first_exporter_failure_is_wrapped_with_cause_and_output_index(
    monkeypatch,
    tmp_path,
):
    original_error = RuntimeError("disk failed")

    def fail_export(*args, **kwargs):
        raise original_error

    monkeypatch.setattr(actions_module, "save_output", fail_export)

    with pytest.raises(OutputExecutionError) as captured:
        execute_structured_result(make_result(output()), tmp_path)

    assert captured.value.field_path == "outputs[0]"
    assert captured.value.__cause__ is original_error
    assert captured.value.details == {"path": str(tmp_path / "result.txt")}


def test_second_exporter_failure_can_leave_first_successful_file(
    monkeypatch,
    tmp_path,
):
    original_save_output = actions_module.save_output
    original_error = RuntimeError("second failed")

    def fail_second(content, output_path, title=None):
        if output_path.name == "second.txt":
            raise original_error
        return original_save_output(content, output_path, title=title)

    monkeypatch.setattr(actions_module, "save_output", fail_second)
    result = make_result(
        output("txt", "first.txt", "First"),
        output("txt", "second.txt", "Second"),
    )

    with pytest.raises(OutputExecutionError) as captured:
        execute_structured_result(result, tmp_path)

    assert captured.value.field_path == "outputs[1]"
    assert captured.value.__cause__ is original_error
    assert (tmp_path / "first.txt").read_text(encoding="utf-8") == "First"
    assert not (tmp_path / "second.txt").exists()


def test_xlsx_invalid_and_duplicate_names_are_executed_from_plan(tmp_path):
    tables = [
        ResultTable(name="Financeiro/2026", rows=[["1"]]),
        ResultTable(name="Data", rows=[["2"]]),
        ResultTable(name="data", rows=[["3"]]),
        ResultTable(name="Data", rows=[["4"]]),
    ]
    created = execute_structured_result(
        make_result(output("xlsx", "sheets.xlsx", tables=tables)),
        tmp_path,
    )[0]

    workbook = load_workbook(created)
    assert workbook.sheetnames == ["Financeiro_2026", "Data", "data_2", "Data_3"]
    workbook.close()
    assert [table.name for table in tables] == [
        "Financeiro/2026",
        "Data",
        "data",
        "Data",
    ]


def test_xlsx_planned_truncation_is_used_without_mutating_original_table(tmp_path):
    original_name = "A" * 40
    table = ResultTable(name=original_name, rows=[["value"]])
    created = execute_structured_result(
        make_result(output("xlsx", "long.xlsx", tables=[table])),
        tmp_path,
    )[0]

    workbook = load_workbook(created)
    assert workbook.sheetnames == ["A" * 31]
    workbook.close()
    assert table.name == original_name
