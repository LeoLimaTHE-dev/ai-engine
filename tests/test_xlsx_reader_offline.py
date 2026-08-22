from openpyxl import Workbook

from ai_engine.readers.xlsx_reader import read_xlsx


def test_read_xlsx_produces_one_table_per_worksheet(tmp_path):
    input_path = tmp_path / "multiple_worksheets.xlsx"

    workbook = Workbook()
    first = workbook.active
    first.title = "First"
    first.append(["name", "value"])
    first.append(["alpha", 10])

    second = workbook.create_sheet("Second")
    second.append(["item", "status"])
    second.append(["beta", "ok"])

    workbook.save(input_path)

    document = read_xlsx(input_path)

    assert len(document.tables) == 2
    assert document.tables[0].name == "First"
    assert document.tables[0].rows == [
        ["name", "value"],
        ["alpha", "10"],
    ]
    assert document.tables[1].name == "Second"
    assert document.tables[1].rows == [
        ["item", "status"],
        ["beta", "ok"],
    ]


def test_read_xlsx_represents_empty_worksheet_with_empty_table(tmp_path):
    input_path = tmp_path / "empty_worksheet.xlsx"

    workbook = Workbook()
    workbook.active.title = "Empty"
    workbook.save(input_path)

    document = read_xlsx(input_path)

    assert len(document.tables) == 1
    assert document.tables[0].name == "Empty"
    assert document.tables[0].rows == []

