import pymupdf
import pytest
from docx import Document
from openpyxl import load_workbook

from ai_engine.exporters import (
    save_docx,
    save_output,
    save_pdf,
    save_text,
    save_xlsx,
    save_xlsx_tables,
)
from ai_engine.results import ResultTable


@pytest.mark.parametrize("extension", [".txt", ".md"])
def test_save_text_preserves_content(extension, tmp_path):
    path = tmp_path / f"content{extension}"
    content = "Título\n\nConteúdo com acentuação."

    result = save_text(content, path)

    assert result == path
    assert path.read_text(encoding="utf-8") == content


def test_save_docx_creates_reopenable_document_with_text(tmp_path):
    path = tmp_path / "report.docx"

    result = save_docx(
        content="Document body",
        output_path=path,
        title="Document title",
    )

    assert result == path
    reopened = Document(path)
    assert [paragraph.text for paragraph in reopened.paragraphs] == [
        "Document title",
        "Document body",
    ]


def test_save_pdf_creates_nonempty_pdf_with_extractable_text(tmp_path):
    path = tmp_path / "report.pdf"

    result = save_pdf(
        content="First paragraph\nSecond paragraph",
        output_path=path,
        title="PDF title",
    )

    assert result == path
    assert path.stat().st_size > 0
    reopened = pymupdf.open(path)
    extracted = "\n".join(page.get_text() for page in reopened)
    reopened.close()
    assert "PDF title" in extracted
    assert "First paragraph" in extracted
    assert "Second paragraph" in extracted


def test_save_xlsx_creates_expected_worksheet_and_text(tmp_path):
    path = tmp_path / "result.xlsx"

    result = save_xlsx(
        content="First line\nSecond line",
        output_path=path,
        title="Workbook title",
    )

    assert result == path
    workbook = load_workbook(path)
    worksheet = workbook["Resultado"]
    assert worksheet["A1"].value == "Workbook title"
    assert worksheet["A3"].value == "First line"
    assert worksheet["A4"].value == "Second line"
    workbook.close()


def test_save_xlsx_tables_creates_sheet_per_table_with_headers_and_rows(tmp_path):
    path = tmp_path / "tables.xlsx"
    tables = [
        ResultTable(
            name="People",
            headers=["name", "city"],
            rows=[["Alice", "São Paulo"], ["Bob", "Recife"]],
        ),
        ResultTable(
            name="Totals",
            headers=["item", "amount"],
            rows=[["A", "10"]],
        ),
    ]

    result = save_xlsx_tables(tables, path)

    assert result == path
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["People", "Totals"]
    people = workbook["People"]
    assert people["A1"].value == "name"
    assert people["B1"].value == "city"
    assert people["A2"].value == "Alice"
    assert people["B2"].value == "São Paulo"
    assert people["A3"].value == "Bob"
    totals = workbook["Totals"]
    assert totals["A1"].value == "item"
    assert totals["B1"].value == "amount"
    assert totals["A2"].value == "A"
    assert totals["B2"].value == "10"
    workbook.close()


@pytest.mark.parametrize("extension", [".txt", ".md", ".docx", ".pdf", ".xlsx"])
def test_save_output_dispatches_each_supported_format(extension, tmp_path):
    path = tmp_path / f"output{extension}"

    result = save_output(
        content="Exported content",
        output_path=path,
        title="Export title",
    )

    assert result == path
    assert path.exists()
    assert path.stat().st_size > 0


def test_save_output_rejects_unsupported_format(tmp_path):
    path = tmp_path / "output.csv"

    with pytest.raises(ValueError, match="Unsupported output format: .csv"):
        save_output("content", path)
